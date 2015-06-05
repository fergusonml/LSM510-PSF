# coding: utf-8
from openpyxl import load_workbook
from openpyxl.cell import get_column_letter
import numpy as np
#from numarray import *
from array import *
#from numeric import *

def read_xlsx2(filename):
 Data_Only=True
 a=[]
 b=[]
 #wb=load_workbook('200nmbeadlsmpsfvalues.xlsx',data_only=True)
 wb=load_workbook(filename,data_only=True)
 #print wb.get_sheet_names()
 ws=wb['Sheet1']
 #a=ws['A1':'A197']
 for col_idx in xrange(4, 6):
  col = get_column_letter(col_idx)
  for row in xrange(1, 11):
   #ws.cell(’%s%s’%(col, row)).value = ’%s%s’ % (col, row)
   a.append(ws.cell("%s%s" % (col, row)).value)
   #print ws.cell(col, row).value
   b.extend(a[10:])
   a=a[:10]
 #print a[:196],a[196:]
 #print [a,b]
 return [a,b]

